# -*-coding: utf8 -*-

from openpyxl import load_workbook

class ExcelExtractor:
    workbook = None

    def __init__(self, file):
        self.workbook = load_workbook(filename = file)
        sheets = self.workbook.get_sheet_names()
        for sheet in sheets:
            ws = self.workbook.get_sheet_by_name(sheet)
            print(ws['A2'].value)